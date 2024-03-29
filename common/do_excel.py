#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# Author:likangming
# Created time:2019/4/17 16:41
# Filename:do_excel.py
import openpyxl
# from ALA_WeChat_1.common import http_request


class Case:
    """
    测试用例集，每一个测试用例，实际上就是它的一个实例
    """

    def __init__(self):  # 每个类属性对应excel里的一个字段，以自己的为主
        self.case_id = None
        self.title = None
        self.url = None
        self.data = None
        self.method = None
        self.header = None
        self.expected = None
        self.actual = None
        self.result = None
        self.check_sql = None
        self.sql_data = None


class DoExcel:

    def __init__(self, file_name, sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name
        self.workbook = openpyxl.load_workbook(file_name)  # 读取excel文件
        self.sheet = self.workbook[sheet_name]  # 指定对应的sheet

    def get_case(self):  # 根据测试用例集的类属性获取excel的对应数据
        max_row = self.sheet.max_row  # 获取最大行数
        cases = []
        for i in range(2, max_row + 1):  # 从第二行开始读取数据
            case = Case()  # 实例
            case.case_id = self.sheet.cell(row=i, column=1).value
            case.title = self.sheet.cell(row=i, column=2).value
            case.url = self.sheet.cell(row=i, column=3).value
            case.data = self.sheet.cell(row=i, column=4).value
            case.method = self.sheet.cell(row=i, column=5).value
            case.expected = self.sheet.cell(row=i, column=6).value
            case.check_sql = self.sheet.cell(row=i, column=9).value
            cases.append(case)
        self.workbook.close()  # 关闭workbook
        return cases  # 返回case列表

    def writer_result(self, row, actual, result, sql_data=None):  # 写入数据到excel中
        sheet = self.workbook[self.sheet_name]
        sheet.cell(row, 7).value = actual
        sheet.cell(row, 8).value = result
        sheet.cell(row, 10).value = sql_data
        self.workbook.save(filename=self.file_name)  # 保存
        self.workbook.close()
